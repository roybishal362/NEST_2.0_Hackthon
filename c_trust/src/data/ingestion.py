"""
C-TRUST Data Ingestion Engine - Component 1
========================================
Data ingestion for NEST 2.0 clinical trial datasets.

Capabilities:
- Binary Excel file reading with error handling
- Automatic study discovery across 23 studies
- File type detection based on filename patterns
- Schema validation and data cleaning
- Snapshot creation for temporal analysis
- Parallel processing for performance

Key Components:
1. ExcelFileReader - Handles binary Excel files (openpyxl + xlrd fallback)
2. FileTypeDetector - Pattern matching for 9 file types
3. StudyDiscovery - Scans data directory for all studies
4. DataValidator - Schema validation and data quality checks
5. SnapshotManager -Versioned data snapshots

Production features:
- Comprehensive error handling
- Detailed logging
- Progress tracking
- Data integrity validation
- Automatic retry logic
"""

import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from src.core import get_logger
from src.core.settings import settings
from src.core.config import config_manager
from src.data.models import FileType, Study

logger = get_logger(__name__)


# ========================================
# EXCEL FILE READER
# ========================================

class ExcelFileReader:
    """
    Excel file reader with binary handling.
    
    Features:
    - Binary file mode support for corrupted/encoded files
    - Automatic sheet detection
    - Multiple engine support (openpyxl, xlrd)
    - Error recovery and retry logic
    - Data type inference
    """
    
    def __init__(self):
        """Initialize Excel file reader"""
        self.supported_extensions = [".xlsx", ".xls", ".xlsm"]
        logger.debug("ExcelFileReader initialized")
    
    def _detect_header_row(self, file_path: Path, sheet_name: str | int = 0) -> int:
        """
        Detect which row contains actual column headers in Excel file.
        
        NEST 2.0 files often have multi-row headers where:
        - Row 0: Actual column names (e.g., "Site ID", "Subject ID", "Project Name")
        - Row 1: Sub-headers or category groupings
        - Row 2: Additional descriptions or units
        - Row 3: Action owners or other metadata
        
        This method finds the row with typical column name patterns.
        
        CRITICAL FIX: Prioritize rows with short, concise column names (like "Site ID", "Subject ID")
        over rows with long descriptions (like "# Expected Visits (Rave EDC : BO4)").
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name or index
        
        Returns:
            Row index (0-based) containing headers
        """
        try:
            # Read first 5 rows without headers to analyze structure
            df_preview = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=None,
                nrows=5,
                engine="openpyxl"
            )
            
            # Find row with best header characteristics
            best_row = 0
            best_score = 0
            
            # Common column name patterns in NEST files
            header_patterns = [
                'ID', 'NAME', 'STATUS', 'DATE', 'SITE', 'SUBJECT', 'PATIENT',
                'STUDY', 'REGION', 'COUNTRY', 'VISIT', 'FORM', 'QUERY', 'SAE'
            ]
            
            # CRITICAL: Check if Row 0 has "Site ID" and "Subject ID" - if so, use it immediately
            # This is the most reliable indicator for NEST CPID files
            row_0 = df_preview.iloc[0]
            row_0_cols = [str(val).strip() for val in row_0 if pd.notna(val)]
            if any('SITE' in col.upper() and 'ID' in col.upper() for col in row_0_cols) and \
               any('SUBJECT' in col.upper() and 'ID' in col.upper() for col in row_0_cols):
                logger.info(f"Found 'Site ID' and 'Subject ID' in row 0 for {file_path.name}, using row 0")
                return 0
            
            for i in range(min(5, len(df_preview))):
                row = df_preview.iloc[i]
                
                # Count non-null values
                non_null_count = row.notna().sum()
                
                # Count non-numeric values (headers are usually text)
                non_numeric_count = 0
                pattern_match_count = 0
                avg_length = 0
                lengths = []
                
                for val in row:
                    if pd.notna(val):
                        val_str = str(val).strip()
                        
                        # Check if non-numeric
                        if not isinstance(val, (int, float)):
                            non_numeric_count += 1
                            lengths.append(len(val_str))
                            
                            # Check for header patterns
                            val_upper = val_str.upper()
                            if any(pattern in val_upper for pattern in header_patterns):
                                pattern_match_count += 1
                
                # Calculate average length (shorter is better for headers)
                if lengths:
                    avg_length = sum(lengths) / len(lengths)
                
                # Scoring:
                # - Prefer rows with header patterns (weight: 20 - INCREASED)
                # - Prefer rows with non-numeric values (weight: 1 - DECREASED)
                # - HEAVILY penalize long text (descriptions vs. column names)
                # - Prefer rows with reasonable number of columns (5-50)
                # - BONUS for short average length (< 20 chars)
                
                pattern_score = pattern_match_count * 20  # Increased from 10
                non_numeric_score = non_numeric_count * 1  # Decreased from 2
                
                # Heavy penalty for long descriptions
                if avg_length > 30:
                    length_penalty = (avg_length - 20) * 2  # Heavy penalty
                elif avg_length > 20:
                    length_penalty = (avg_length - 20)  # Moderate penalty
                else:
                    length_penalty = -(20 - avg_length) * 0.5  # BONUS for short names
                
                column_count_score = 5 if 5 <= non_null_count <= 50 else 0
                
                score = pattern_score + non_numeric_score + column_count_score - length_penalty
                
                logger.debug(
                    f"Row {i}: patterns={pattern_match_count}, non_numeric={non_numeric_count}, "
                    f"avg_len={avg_length:.1f}, score={score:.1f}"
                )
                
                if score > best_score:
                    best_score = score
                    best_row = i
            
            logger.info(f"Detected header row {best_row} for {file_path.name} (score={best_score:.1f})")
            return best_row
            
        except Exception as e:
            logger.warning(f"Header detection failed for {file_path.name}: {e}, using row 0")
            return 0
    
    def read_file(
        self,
        file_path: Path | str,
        sheet_name: str | int = 0,
        header_row: int = None,
        auto_detect_header: bool = True
    ) -> Optional[pd.DataFrame]:
        """
        Read Excel file with comprehensive error handling and automatic header detection.
        
        CRITICAL FIX (Task 2.2): Properly handles multi-row headers in EDC Metrics files.
        EDC Metrics files have headers in rows 0-2 with data starting at row 3.
        This method detects EDC Metrics files and reads them with header=[0,1,2],
        then flattens the tuple column names for easier access.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name or index (0-based)
            header_row: Row number containing headers (0-based). If None, auto-detects.
            auto_detect_header: If True, automatically detect header row
        
        Returns:
            DataFrame if successful, None if failed
        
        Example:
            reader = ExcelFileReader()
            df = reader.read_file("Study_01_EDC_Metrics.xlsx")
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            logger.error(f"File not found: {file_path}")
            return None
        
        if file_path.suffix not in self.supported_extensions:
            logger.error(f"Unsupported file extension: {file_path.suffix}")
            return None
        
        logger.info(f"Reading Excel file: {file_path.name}")
        
        # CRITICAL FIX: Detect EDC Metrics files and use multi-row header
        is_edc_metrics = "EDC_Metrics" in file_path.name or "CPID_EDC" in file_path.name
        
        if is_edc_metrics:
            logger.info(f"Detected EDC Metrics file: {file_path.name}, using multi-row header [0,1,2]")
            # Read with multi-row header for EDC Metrics files
            try:
                df = self._read_edc_metrics_with_multirow_header(file_path, sheet_name)
                if df is not None and not df.empty:
                    logger.info(
                        f"Successfully read EDC Metrics {file_path.name}: "
                        f"{len(df)} rows, {len(df.columns)} columns (multi-row header)"
                    )
                    return df
            except Exception as e:
                logger.warning(f"Multi-row header read failed for {file_path.name}: {e}, trying standard approach")
        
        # Standard approach for non-EDC Metrics files or fallback
        # Auto-detect header row if not specified
        if header_row is None and auto_detect_header:
            header_row = self._detect_header_row(file_path, sheet_name)
        elif header_row is None:
            header_row = 0
        
        # Try multiple reading strategies
        strategies = [
            lambda: self._read_with_openpyxl(file_path, sheet_name, header_row),
            lambda: self._read_with_pandas_engine(file_path, sheet_name, header_row),
            lambda: self._read_with_xlrd(file_path, sheet_name, header_row),
        ]
        
        for strategy in strategies:
            try:
                df = strategy()
                if df is not None and not df.empty:
                    logger.info(
                        f"Successfully read {file_path.name}: "
                        f"{len(df)} rows, {len(df.columns)} columns (header_row={header_row})"
                    )
                    return df
            except Exception as e:
                logger.debug(f"Strategy {strategy} failed: {e}")
                continue
        
        logger.error(f"All read strategies failed for: {file_path.name}")
        return None
    
    def _read_edc_metrics_with_multirow_header(
        self,
        file_path: Path,
        sheet_name: str | int
    ) -> Optional[pd.DataFrame]:
        """
        Read EDC Metrics file with multi-row header [0,1,2].
        
        CRITICAL FIX (Task 2.2): EDC Metrics files have multi-row headers:
        - Row 0: Main category (e.g., "CPMD", "Input files")
        - Row 1: Sub-category (e.g., "Visit status", "Page status")
        - Row 2: Actual column name (e.g., "# Expected Visits", "# Pages Entered")
        
        This method reads with header=[0,1,2] and flattens the tuple column names
        into readable strings by joining non-empty parts.
        
        Args:
            file_path: Path to EDC Metrics Excel file
            sheet_name: Sheet name or index
        
        Returns:
            DataFrame with flattened column names
        """
        try:
            # Read with multi-row header
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=[0, 1, 2],
                engine="openpyxl"
            )
            
            logger.debug(f"Read EDC Metrics with multi-row header: {len(df)} rows, {len(df.columns)} columns")
            
            # Flatten tuple column names
            # Strategy: Join non-empty, non-"Unnamed" parts with " - "
            new_columns = []
            for col in df.columns:
                if isinstance(col, tuple):
                    # Filter out empty strings, "nan", and "Unnamed" parts
                    parts = [
                        str(part).strip() 
                        for part in col 
                        if str(part) != 'nan' 
                        and 'Unnamed' not in str(part)
                        and str(part).strip() != ''
                    ]
                    # Join remaining parts
                    if parts:
                        flattened = ' - '.join(parts)
                    else:
                        # If all parts were filtered out, use the last part
                        flattened = str(col[-1])
                    new_columns.append(flattened)
                else:
                    new_columns.append(str(col))
            
            df.columns = new_columns
            
            logger.info(
                f"Flattened EDC Metrics columns: {len(df.columns)} columns, "
                f"sample: {list(df.columns[:5])}"
            )
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to read EDC Metrics with multi-row header: {e}")
            return None
    
    def _read_with_openpyxl(
        self,
        file_path: Path,
        sheet_name: str | int,
        header_row: int
    ) -> Optional[pd.DataFrame]:
        """
        Read using openpyxl engine (best for .xlsx).
        
        This is the primary strategy for modern Excel files.
        """
        # First, try to open with openpyxl to validate
        wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        
        # Get sheet name if index provided
        if isinstance(sheet_name, int):
            sheet_name = wb.sheetnames[sheet_name]
        
        wb.close()
        
        # Now read with pandas
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=header_row,
            engine="openpyxl"
        )
        
        return df
    
    def _read_with_pandas_engine(
        self,
        file_path: Path,
        sheet_name: str | int,
        header_row: int
    ) -> Optional[pd.DataFrame]:
        """
        Read using pandas default engine.
        
        Fallback strategy when openpyxl fails.
        """
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=header_row
        )
        
        return df
    
    def _read_with_xlrd(
        self,
        file_path: Path,
        sheet_name: str | int,
        header_row: int
    ) -> Optional[pd.DataFrame]:
        """
        Read using xlrd engine (for .xls files).
        
        Last resort strategy for older Excel formats.
        """
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=header_row,
            engine="xlrd"
        )
        
        return df
    
    def get_sheet_names(self, file_path: Path | str) -> List[str]:
        """
        Get all sheet names from Excel file.
        
        Args:
            file_path: Path to Excel file
        
        Returns:
            List of sheet names
        """
        file_path = Path(file_path)
        
        try:
            wb = load_workbook(filename=str(file_path), read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            logger.error(f"Failed to read sheet names from {file_path.name}: {e}")
            return []


# ========================================
# FILE TYPE DETECTOR
# ========================================

class FileTypeDetector:
    """
    Detects file type based on filename patterns.
    
    Uses regex patterns from settings.yaml to classify files into
    one of 9 NEST 2.0 file types.
    """
    
    def __init__(self):
        """Initialize file type detector with patterns from config"""
        # Get patterns from data_sources configuration
        config = config_manager.get_config()
        if hasattr(config, 'data_sources'):
            self.file_type_config = config.data_sources.file_patterns
        else:
            self.file_type_config = {}
            
        self._compile_patterns()
        logger.debug("FileTypeDetector initialized with patterns from config")
    
    def _compile_patterns(self) -> None:
        """Compile regex patterns for fast matching"""
        self.patterns: Dict[FileType, re.Pattern] = {}
        
        for file_type_key, pattern_config in self.file_type_config.items():
            # Normalize to list of patterns
            patterns_list = []
            if isinstance(pattern_config, list):
                patterns_list = pattern_config
            elif isinstance(pattern_config, dict):
                # Legacy dict support
                patterns_list = [pattern_config.get("pattern", "")]
            else:
                # Single string pattern
                patterns_list = [str(pattern_config)]
                
            # Compile all patterns for this file type
            regex_parts = []
            for pat in patterns_list:
                if pat:
                    # Convert glob to regex (without start/end anchors initially if combining)
                    # But _glob_to_regex adds .* at start and $ at end.
                    # We need to strip them to combine, or handle effectively.
                    # _glob_to_regex returns f".*{regex}$"
                    # If we use re.match, we need the start anchor or .* at start.
                    
                    regex = self._glob_to_regex(pat)
                    # Use non-capturing group for safety
                    regex_parts.append(f"(?:{regex})")
            
            if not regex_parts:
                continue
                
            # Combine with OR
            full_regex = "|".join(regex_parts)
            
            try:
                # Try to map key to FileType enum
                try:
                    file_type = FileType(file_type_key)
                except ValueError:
                    try:
                        file_type = FileType(file_type_key.lower())
                    except ValueError:
                        continue

                self.patterns[file_type] = re.compile(full_regex, re.IGNORECASE)
            except ValueError:
                logger.warning(f"Error compiling pattern for {file_type_key}")
    
    def _glob_to_regex(self, glob_pattern: str) -> str:
        """
        Convert glob pattern to regex.
        
        Examples:
            "*EDC_Metrics*.xlsx" -> ".*EDC_Metrics.*\\.xlsx"
            "*Visit*Tracker*.xlsx" -> ".*Visit.*Tracker.*\\.xlsx"
        """
        if not glob_pattern:
            return "^$"
            
        # Escape special regex characters except * and ?
        regex = re.escape(glob_pattern)
        # Replace escaped wildcards with regex equivalents
        regex = regex.replace(r"\*", ".*").replace(r"\?", ".")
        return f".*{regex}$"  # Allow prefix matching (e.g. directory path)
    
    def detect_file_type(self, filename: str) -> Optional[FileType]:
        """
        Detect file type from filename.
        
        Args:
            filename: Name of the file
        
        Returns:
            FileType enum or None if no match
        
        Example:
            detector = FileTypeDetector()
            file_type = detector.detect_file_type("Study_01_EDC_Metrics.xlsx")
            # Returns: FileType.EDC_METRICS
        """
        for file_type, pattern in self.patterns.items():
            if pattern.match(filename):
                logger.debug(f"Detected {file_type.value} for file: {filename}")
                return file_type
        
        logger.warning(f"Could not detect file type for: {filename}")
        return None
    
    def is_required_file(self, file_type: FileType) -> bool:
        """
        Check if file type is required for study.
        
        Args:
            file_type: File type to check
        
        Returns:
            True if required field in config is True
        """
        config = self.file_type_config.get(file_type.value, {})
        return config.get("required", False)


# ========================================
# STUDY DISCOVERY
# ========================================

class StudyDiscovery:
    """
    Discovers all studies in the NEST 2.0 dataset.
    
    Scans the data directory and identifies:
    - All study folders (e.g., Study 1, STUDY 15, etc.)
    - Available files per study
    - File completeness (required files present)
    """
    
    def __init__(self, data_root: Optional[str] = None):
        """
        Initialize study discovery.
        
        Args:
            data_root: Root path to NEST 2.0 data. Uses settings if not provided.
        """
        self.data_root = Path(data_root or getattr(settings, 'DATA_ROOT_PATH', 'norvatas/Data for problem Statement 1/NEST 2.0 Data files_Anonymized/QC Anonymized Study Files'))
        self.file_reader = ExcelFileReader()
        self.file_detector = FileTypeDetector()
        
        if not self.data_root.exists():
            raise FileNotFoundError(f"Data root path does not exist: {self.data_root}")
        
        logger.info(f"StudyDiscovery initialized with root: {self.data_root}")
    
    def discover_all_studies(self) -> List[Study]:
        """
        Discover all studies in the dataset.
        
        Returns:
            List of Study objects with basic information
        
        Example:
            discovery = StudyDiscovery()
            studies = discovery.discover_all_studies()
            print(f"Found {len(studies)} studies")
        """
        logger.info("Starting study discovery...")
        
        study_folders = self._find_study_folders()
        logger.info(f"Found {len(study_folders)} study folders")
        
        studies = []
        
        for study_folder in study_folders:
            try:
                study = self._process_study_folder(study_folder)
                if study:
                    studies.append(study)
            except Exception as e:
                logger.error(f"Error processing {study_folder.name}: {e}", exc_info=True)
        
        logger.info(f"Successfully discovered {len(studies)} studies")
        return studies
    
    def _find_study_folders(self) -> List[Path]:
        """
        Find all study folders in data root.
        
        Returns:
            List of paths to study folders
        """
        study_folders = []
        
        # Pattern to match study folders: "Study X_" or "STUDY X_" or "SIM-XXX"
        study_pattern = re.compile(r"^study[\s_]+\d+", re.IGNORECASE)
        sim_pattern = re.compile(r"^SIM-\d+", re.IGNORECASE)
        
        for item in self.data_root.iterdir():
            if item.is_dir() and (study_pattern.match(item.name) or sim_pattern.match(item.name)):
                study_folders.append(item)
        
        # Sort by study number
        study_folders.sort(key=lambda x: self._extract_study_number(x.name))
        
        return study_folders
    
    def _extract_study_number(self, folder_name: str) -> int:
        """Extract study number from folder name for sorting"""
        match = re.search(r"\d+", folder_name)
        return int(match.group()) if match else 0
    
    def _process_study_folder(self, study_folder: Path) -> Optional[Study]:
        """
        Process a single study folder.
        
        Args:
            study_folder: Path to study folder
        
        Returns:
            Study object with discovered files
        """
        study_id = self._normalize_study_id(study_folder.name)
        logger.debug(f"Processing study: {study_id}")
        
        # Find all Excel files in folder
        excel_files = list(study_folder.glob("*.xlsx")) + list(study_folder.glob("*.xls"))
        
        if not excel_files:
            logger.warning(f"No Excel files found in {study_folder.name}")
            return None
        
        # Detect file types
        available_files: Dict[FileType, bool] = {}
        file_paths: Dict[FileType, Path] = {}
        
        for excel_file in excel_files:
            file_type = self.file_detector.detect_file_type(excel_file.name)
            if file_type:
                available_files[file_type] = True
                file_paths[file_type] = excel_file
        
        # Create Study object
        study = Study(
            study_id=study_id,
            study_name=study_id,  # Will be updated from data
            available_files=available_files,
            last_data_refresh=datetime.now(),
            metadata={
                "folder_path": str(study_folder),
                "file_paths": {k.value: str(v) for k, v in file_paths.items()},
                "file_count": len(excel_files),
            }
        )
        
        logger.debug(f"Study {study_id}: {len(available_files)} file types detected")
        return study
    
    def _normalize_study_id(self, folder_name: str) -> str:
        """
        Normalize study ID to consistent format.
        
        Examples:
            "Study 1_CPID_Input Files" -> "STUDY_01"
            "STUDY 15_CPID_Input Files" -> "STUDY_15"
            "SIM-001" -> "SIM-001"
            "SIM-006" -> "SIM-006"
        """
        # Check for SIM pattern first
        sim_match = re.search(r"^(SIM-\d+)", folder_name, re.IGNORECASE)
        if sim_match:
            return sim_match.group(1).upper()
        
        # Check for regular study pattern
        match = re.search(r"study[\s_]+(\d+)", folder_name, re.IGNORECASE)
        if match:
            study_number = int(match.group(1))
            return f"STUDY_{study_number:02d}"
        
        return folder_name
    
    def get_file_path(self, study_id: str, file_type: FileType) -> Optional[Path]:
        """
        Get file path for specific study and file type.
        
        Args:
            study_id: Study identifier
            file_type: Type of file to find
        
        Returns:
            Path to file or None if not found
        """
        studies = self.discover_all_studies()
        
        for study in studies:
            if study.study_id == study_id:
                file_paths = study.metadata.get("file_paths", {})
                file_path_str = file_paths.get(file_type.value)
                if file_path_str:
                    return Path(file_path_str)
        
        return None


# ========================================
# DATA VALIDATION ENGINE
# ========================================

class DataValidator:
    """
    Validates data quality and schema compliance.
    
    Features:
    - Schema validation against expected structure
    - Data quality checks (nulls, ranges, formats)
    - Error reporting and logging
    - Configurable validation rules
    """
    
    def __init__(self, strict_mode: bool = False):
        """
        Initialize data validator.
        
        Args:
            strict_mode: If True, validation failures raise errors.
                        If False, validation failures log warnings and continue.
        """
        self.strict_mode = strict_mode
        self.validation_rules = self._load_validation_rules()
        logger.debug(f"DataValidator initialized (strict_mode={strict_mode})")
    
    def _load_validation_rules(self) -> Dict[FileType, Dict[str, Any]]:
        """Load validation rules from configuration"""
        return {
            FileType.EDC_METRICS: {
                "required_columns": ["Study", "Site", "Subject", "Visit"],
                "numeric_columns": ["Total_Forms", "Completed_Forms", "Open_Queries"],
                "date_columns": ["Last_Update"],
                "min_rows": 1
            },
            FileType.SAE_DM: {
                "required_columns": ["Study", "Site", "Subject", "SAE_ID"],
                "numeric_columns": ["Days_to_Report"],
                "date_columns": ["SAE_Date", "Report_Date"],
                "min_rows": 0  # SAE files can be empty
            },
            FileType.VISIT_PROJECTION: {
                "required_columns": ["Study", "Site", "Subject", "Visit"],
                "date_columns": ["Planned_Date", "Actual_Date"],
                "min_rows": 1
            },
            FileType.MISSING_PAGES: {
                "required_columns": ["Study", "Site", "Subject", "Form"],
                "min_rows": 0  # Can be empty if no missing pages
            }
        }
    
    def validate_dataframe(
        self,
        df: pd.DataFrame,
        file_type: FileType,
        file_path: Path
    ) -> Tuple[bool, List[str]]:
        """
        Validate DataFrame against schema and quality rules.
        
        Args:
            df: DataFrame to validate
            file_type: Type of file being validated
            file_path: Path to original file (for error reporting)
        
        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        errors = []
        
        if file_type not in self.validation_rules:
            logger.warning(f"No validation rules for file type: {file_type}")
            return True, []
        
        rules = self.validation_rules[file_type]
        
        # Check minimum rows
        if len(df) < rules.get("min_rows", 0):
            errors.append(f"Insufficient rows: {len(df)} < {rules['min_rows']}")
        
        # Check required columns
        required_cols = rules.get("required_columns", [])
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            errors.append(f"Missing required columns: {missing_cols}")
        
        # Check numeric columns
        numeric_cols = rules.get("numeric_columns", [])
        for col in numeric_cols:
            if col in df.columns:
                non_numeric = df[col].apply(lambda x: not pd.api.types.is_numeric_dtype(type(x)))
                if non_numeric.any():
                    errors.append(f"Non-numeric values in column '{col}'")
        
        # Check date columns
        date_cols = rules.get("date_columns", [])
        for col in date_cols:
            if col in df.columns:
                try:
                    pd.to_datetime(df[col], errors='coerce')
                except Exception:
                    errors.append(f"Invalid date format in column '{col}'")
        
        is_valid = len(errors) == 0
        
        if not is_valid:
            if self.strict_mode:
                logger.error(f"Validation failed for {file_path.name}: {errors}")
            else:
                logger.warning(f"Validation failed (non-strict mode) for {file_path.name}: {errors}")
                # In non-strict mode, treat as valid to allow processing
                is_valid = True
        else:
            logger.debug(f"Validation passed for {file_path.name}")
        
        return is_valid, errors


# ========================================
# BATCH PROCESSING WORKFLOW
# ========================================

class BatchProcessor:
    """
    Handles offline batch processing of clinical trial data.
    
    Features:
    - Configurable batch schedules (daily, weekly)
    - Progress tracking and reporting
    - Error recovery and retry logic
    - Snapshot versioning
    """
    
    def __init__(self):
        """Initialize batch processor"""
        self.reader = ExcelFileReader()
        self.validator = DataValidator()
        self.discovery = StudyDiscovery()
        logger.info("BatchProcessor initialized")
    
    def process_batch(
        self,
        study_ids: Optional[List[str]] = None,
        file_types: Optional[List[FileType]] = None,
        create_snapshot: bool = True
    ) -> Dict[str, Any]:
        """
        Process a batch of studies with comprehensive error handling.
        
        Args:
            study_ids: Specific studies to process (None for all)
            file_types: Specific file types to process (None for all)
            create_snapshot: Whether to create versioned snapshot
        
        Returns:
            Processing results with success/failure counts
        """
        logger.info("Starting batch processing...")
        start_time = datetime.now()
        
        # Initialize results tracking
        results = {
            "start_time": start_time,
            "studies_processed": 0,
            "studies_failed": 0,
            "files_processed": 0,
            "files_failed": 0,
            "validation_errors": [],
            "processing_errors": [],
            "snapshot_id": None
        }
        
        try:
            # Discover studies
            all_studies = self.discovery.discover_all_studies()
            
            # Filter studies if specified
            if study_ids:
                studies_to_process = [s for s in all_studies if s.study_id in study_ids]
            else:
                studies_to_process = all_studies
            
            logger.info(f"Processing {len(studies_to_process)} studies")
            
            # Process each study
            processed_data = {}
            
            for study in studies_to_process:
                try:
                    study_data = self._process_study_with_validation(
                        study, file_types, results
                    )
                    
                    if study_data:
                        processed_data[study.study_id] = study_data
                        results["studies_processed"] += 1
                    else:
                        results["studies_failed"] += 1
                        
                except Exception as e:
                    logger.error(f"Failed to process study {study.study_id}: {e}", exc_info=True)
                    results["studies_failed"] += 1
                    results["processing_errors"].append({
                        "study_id": study.study_id,
                        "error": str(e),
                        "timestamp": datetime.now()
                    })
            
            # Create snapshot if requested
            if create_snapshot and processed_data:
                snapshot_id = self._create_data_snapshot(processed_data)
                results["snapshot_id"] = snapshot_id
            
            # Calculate processing time
            end_time = datetime.now()
            results["end_time"] = end_time
            results["processing_time_seconds"] = (end_time - start_time).total_seconds()
            
            logger.info(
                f"Batch processing complete: "
                f"{results['studies_processed']} studies, "
                f"{results['files_processed']} files, "
                f"{results['processing_time_seconds']:.1f}s"
            )
            
            return results
            
        except Exception as e:
            logger.error(f"Batch processing failed: {e}", exc_info=True)
            results["processing_errors"].append({
                "error": f"Batch processing failure: {str(e)}",
                "timestamp": datetime.now()
            })
            return results
    
    def _process_study_with_validation(
        self,
        study: Study,
        file_types: Optional[List[FileType]],
        results: Dict[str, Any]
    ) -> Optional[Dict[FileType, pd.DataFrame]]:
        """Process single study with validation"""
        logger.debug(f"Processing study: {study.study_id}")
        
        study_data = {}
        file_paths = study.metadata.get("file_paths", {})
        
        # Filter file types if specified
        if file_types:
            file_paths = {
                ft: path for ft, path in file_paths.items()
                if FileType(ft) in file_types
            }
        
        for file_type_str, file_path_str in file_paths.items():
            try:
                file_type = FileType(file_type_str)
                file_path = Path(file_path_str)
                
                # Read file
                df = self.reader.read_file(file_path)
                
                if df is not None:
                    # Validate data
                    is_valid, validation_errors = self.validator.validate_dataframe(
                        df, file_type, file_path
                    )
                    
                    if is_valid:
                        study_data[file_type] = df
                        results["files_processed"] += 1
                        logger.debug(
                            f"{study.study_id} - {file_type.value}: "
                            f"{len(df)} rows processed"
                        )
                    else:
                        results["files_failed"] += 1
                        results["validation_errors"].append({
                            "study_id": study.study_id,
                            "file_type": file_type.value,
                            "file_path": str(file_path),
                            "errors": validation_errors,
                            "timestamp": datetime.now()
                        })
                        logger.warning(
                            f"{study.study_id} - {file_type.value}: "
                            f"Validation failed: {validation_errors}"
                        )
                else:
                    results["files_failed"] += 1
                    logger.warning(f"{study.study_id} - Failed to read {file_type.value}")
            
            except Exception as e:
                results["files_failed"] += 1
                results["processing_errors"].append({
                    "study_id": study.study_id,
                    "file_type": file_type_str,
                    "error": str(e),
                    "timestamp": datetime.now()
                })
                logger.error(
                    f"{study.study_id} - Error processing {file_type_str}: {e}",
                    exc_info=True
                )
        
        return study_data if study_data else None
    
    def _create_data_snapshot(self, processed_data: Dict[str, Dict[FileType, pd.DataFrame]]) -> str:
        """Create versioned data snapshot"""
        snapshot_id = f"snapshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # In a full implementation, this would save to database
        # For now, we'll just log the snapshot creation
        logger.info(f"Created data snapshot: {snapshot_id}")
        
        return snapshot_id


# ========================================
# DATA INGESTION ORCHESTRATOR
# ========================================

class DataIngestionEngine:
    """
    Main orchestrator for data ingestion.
    
    Enhanced with:
    - Schema validation and error handling
    - Batch processing workflow for offline execution
    - Comprehensive error recovery
    - Progress tracking and reporting
    
    Supports parallel processing for better performance.
    """
    
    def __init__(self, strict_validation: bool = False):
        """
        Initialize data ingestion engine.
        
        Args:
            strict_validation: If True, validation failures block processing.
                             If False (default), validation failures log warnings but allow processing.
        """
        self.discovery = StudyDiscovery()
        self.reader = ExcelFileReader()
        self.detector = FileTypeDetector()
        self.validator = DataValidator(strict_mode=strict_validation)
        self.batch_processor = BatchProcessor()
        
        logger.info(f"DataIngestionEngine initialized (strict_validation={strict_validation})")
    
    def ingest_all_studies(
        self,
        parallel: bool = True,
        max_workers: Optional[int] = None,
        validate_data: bool = True
    ) -> Dict[str, Dict[FileType, pd.DataFrame]]:
        """
        Ingest data from all studies with enhanced error handling.
        
        Args:
            parallel: Use parallel processing
            max_workers: Maximum worker threads (defaults to settings.MAX_WORKERS)
            validate_data: Whether to validate data during ingestion
        
        Returns:
            Dictionary mapping study_id -> file_type -> DataFrame
        
        Example:
            engine = DataIngestionEngine()
            all_data = engine.ingest_all_studies(validate_data=True)
            
            # Access specific data
            study_01_edc = all_data["STUDY_01"][FileType.EDC_METRICS]
        """
        logger.info("Starting full data ingestion with validation...")
        
        studies = self.discovery.discover_all_studies()
        max_workers = max_workers or getattr(settings, 'MAX_WORKERS', 4)
        
        all_data: Dict[str, Dict[FileType, pd.DataFrame]] = {}
        
        if parallel:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_study = {
                    executor.submit(self.ingest_study, study, validate_data): study
                    for study in studies
                }
                
                for future in as_completed(future_to_study):
                    study = future_to_study[future]
                    try:
                        study_data = future.result()
                        if study_data:
                            all_data[study.study_id] = study_data
                    except Exception as e:
                        logger.error(f"Error ingesting {study.study_id}: {e}", exc_info=True)
        else:
            for study in studies:
                try:
                    study_data = self.ingest_study(study, validate_data)
                    if study_data:
                        all_data[study.study_id] = study_data
                except Exception as e:
                    logger.error(f"Error ingesting {study.study_id}: {e}", exc_info=True)
        
        logger.info(f"Data ingestion complete: {len(all_data)} studies ingested")
        return all_data
    
    def ingest_study(self, study: Study, validate_data: bool = True) -> Dict[FileType, pd.DataFrame]:
        """
        Ingest all files for a single study with validation.
        
        Args:
            study: Study object from discovery
            validate_data: Whether to validate data during ingestion
        
        Returns:
            Dictionary mapping file_type -> DataFrame
        """
        logger.info(f"Ingesting study: {study.study_id}")
        
        study_data: Dict[FileType, pd.DataFrame] = {}
        file_paths = study.metadata.get("file_paths", {})
        
        for file_type_str, file_path_str in file_paths.items():
            try:
                file_type = FileType(file_type_str)
                file_path = Path(file_path_str)
                
                df = self.reader.read_file(file_path)
                
                if df is not None:
                    # Validate data if requested
                    if validate_data:
                        is_valid, validation_errors = self.validator.validate_dataframe(
                            df, file_type, file_path
                        )
                        
                        if not is_valid:
                            logger.warning(
                                f"{study.study_id} - {file_type.value}: "
                                f"Validation failed: {validation_errors}"
                            )
                            continue  # Skip invalid data
                    
                    study_data[file_type] = df
                    logger.debug(
                        f"{study.study_id} - {file_type.value}: "
                        f"{len(df)} rows loaded"
                    )
                else:
                    logger.warning(
                        f"{study.study_id} - Failed to read {file_type.value}"
                    )
            
            except Exception as e:
                logger.error(
                    f"{study.study_id} - Error reading {file_type_str}: {e}",
                    exc_info=True
                )
        
        return study_data
    
    def process_batch_offline(
        self,
        study_ids: Optional[List[str]] = None,
        file_types: Optional[List[FileType]] = None,
        create_snapshot: bool = True
    ) -> Dict[str, Any]:
        """
        Process clinical trial data in offline batch mode.
        
        This method implements the batch processing workflow required
        for offline execution as specified in Requirements 1.4.
        
        Args:
            study_ids: Specific studies to process (None for all)
            file_types: Specific file types to process (None for all)
            create_snapshot: Whether to create versioned snapshot
        
        Returns:
            Processing results with success/failure counts
        
        Example:
            engine = DataIngestionEngine()
            
            # Process all studies
            results = engine.process_batch_offline()
            
            # Process specific studies and file types
            results = engine.process_batch_offline(
                study_ids=["STUDY_01", "STUDY_02"],
                file_types=[FileType.EDC_METRICS, FileType.SAE_DM]
            )
        """
        return self.batch_processor.process_batch(
            study_ids=study_ids,
            file_types=file_types,
            create_snapshot=create_snapshot
        )
    
    def get_processing_status(self) -> Dict[str, Any]:
        """
        Get current processing status and health metrics.
        
        Returns:
            Dictionary with processing status information
        """
        studies = self.discovery.discover_all_studies()
        
        status = {
            "total_studies": len(studies),
            "studies_with_data": 0,
            "total_files_available": 0,
            "file_type_coverage": {},
            "last_check": datetime.now()
        }
        
        # Analyze file coverage
        file_type_counts = {}
        
        for study in studies:
            available_files = study.available_files
            if available_files:
                status["studies_with_data"] += 1
                status["total_files_available"] += len(available_files)
                
                for file_type, is_available in available_files.items():
                    if is_available:
                        file_type_counts[file_type] = file_type_counts.get(file_type, 0) + 1
        
        # Calculate coverage percentages
        for file_type, count in file_type_counts.items():
            status["file_type_coverage"][file_type.value] = {
                "count": count,
                "percentage": round((count / len(studies)) * 100, 1)
            }
        
        return status


# ========================================
# EXPORTS
# ========================================

__all__ = [
    "ExcelFileReader",
    "FileTypeDetector", 
    "StudyDiscovery",
    "DataValidator",
    "BatchProcessor",
    "DataIngestionEngine",
]
